


def sheetName(excel_file_path,savingPath,savingPath2):
    import os
    from openpyxl import load_workbook
    wb = load_workbook(filename=excel_file_path)

    # Get all sheet names
    ExcelSheets = wb.sheetnames


    with open(savingPath, 'w') as file:
        file.write(f'''{ExcelSheets}''')

    import copy
    dicList = copy.deepcopy(ExcelSheets)
    columns_dict = {str(i+1): dicList[i] for i in range(len(dicList))}
    print(columns_dict)

    with open(savingPath2, 'w') as file:
        file.write(f'''{columns_dict}''')


    return ExcelSheets





def DefaultSplit(savingPath, firstItem):
    import os

    if os.path.exists(savingPath):
        print("file exists")
    else:
        print("file not exists")
        with open(savingPath, 'w') as file:
            file.write(f'''eventLogDiv={firstItem}'''+"\n")
            file.write(f'''otherEntitiesDiv={firstItem}'''+"\n")
            file.write(f'''otherEntitiesRelDiv={firstItem}'''+"\n")
            file.write(f'''activitiesValueDiv={firstItem}'''+"\n")
            file.write(f'''domainDiv={firstItem}'''+"\n")
            file.write(f'''iCDDiv={firstItem}'''+"\n")
            file.write(f'''snomedCtRelNodeDiv={firstItem}'''+"\n")
            file.write(f'''snomedCtRelDiv={firstItem}'''+"\n")
            file.write(f'''dk1Div={firstItem}'''+"\n")
            file.write(f'''dk2Div={firstItem}'''+"\n")
            file.write(f'''dk3Div={firstItem}'''+"\n")
            file.write(f'''dk4Div={firstItem}'''+"\n")
            file.write(f'''dk5Div={firstItem}'''+"\n")
            file.write(f'''dk61Div={firstItem}'''+"\n")
            file.write(f'''dk62Div={firstItem}'''+"\n")
            file.write(f'''dk7Div={firstItem}'''+"\n")


def DefaultValue(file1, file2, save):

    import ast
    with open(file1, 'r') as file:
        data = file.read()
        listData = ast.literal_eval(data)
        #print(listData)

    valueList=[]
    for item in listData:
        #print(item)
        #print(type(item))
        with open(file2, 'r') as file:
            for line in file:
                #print(line)
                variable_name, value = line.split('=')
                variable_name = variable_name.strip()
                #print(variable_name)
                #print(value)
                value_cleaned = value.replace("\n", "")
                if item==variable_name:
                    #print("dsadasd")
                    valueList.append(value_cleaned)

    with open(save, 'w') as file:
        file.write(f'''{valueList}''')

    print(valueList)


