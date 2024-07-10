

def local2_software(excel_file_path,goalList,openingPath1,savingPath1):
    from openpyxl import load_workbook
    wb = load_workbook(filename=excel_file_path)

    # Get all sheet names
    ExcelSheets = wb.sheetnames
    print("sheet_names=",ExcelSheets)


    def get_user_choice(listOrg,prompt_message):
        while True:
            for i, value in enumerate(listOrg, start=1):
                    print(f"{i}. {value}")

            user_input = input(prompt_message)

            if user_input.isdigit():
                user_choice = int(user_input)
                if 1 <= user_choice <= len(listOrg):
                    selected_value = listOrg[user_choice - 1]
                    print(f"You have chosen: {selected_value}")
                    return selected_value
                else:
                    print("Please enter a number within the list's range.")
            else:
                print("Invalid input. Please enter an integer number.")

    EventLog = []
    for variable_to_update in goalList:

        #print(item)
        new_value= get_user_choice(ExcelSheets,f'''Please enter the corresponding column to {variable_to_update}?''')
        print("variable_to_update=", variable_to_update)
        print("new_value=",new_value)

        with open(openingPath1, 'r') as file:
            lines = file.readlines()

        # Update the specific line
        updated = False
        for i, line in enumerate(lines):
            if line.strip().startswith(variable_to_update):
                lines[i] = f"{variable_to_update}={new_value}\n"
                updated = True
                break
        if updated:
            with open(savingPath1, 'w') as file:
                file.writelines(lines)

    return EventLog
